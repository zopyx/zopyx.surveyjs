"""Excel/XLSX converter for Response objects."""

from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font

from .csv_export import CSVWideAdapter, CSVLongAdapter
from .types import Response


def write_xlsx(response: Response, destination: Path, 
               format: str = "wide", max_dynamic_rows: int = 10) -> Path:
    """Write Excel export for a single response.
    
    Args:
        response: Response to export
        destination: Output file path
        format: "wide" or "long"
        max_dynamic_rows: Maximum dynamic rows for wide format
    
    Returns:
        Path to written file
    """
    destination.parent.mkdir(parents=True, exist_ok=True)
    
    if format == "wide":
        adapter = CSVWideAdapter(max_dynamic_rows)
    else:
        adapter = CSVLongAdapter()
    
    rows = adapter.export(response)
    headers = adapter.get_headers([response])
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Survey"
    
    # Write headers with bold font
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
    
    # Write data
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
    
    wb.save(destination)
    return destination


def write_xlsx_multi(responses: List[Response], destination: Path,
                    format: str = "wide", max_dynamic_rows: int = 10) -> Path:
    """Write Excel export for multiple responses.
    
    Args:
        responses: List of responses to export
        destination: Output file path
        format: "wide" or "long"
        max_dynamic_rows: Maximum dynamic rows for wide format
    
    Returns:
        Path to written file
    """
    destination.parent.mkdir(parents=True, exist_ok=True)
    
    if format == "wide":
        adapter = CSVWideAdapter(max_dynamic_rows)
    else:
        adapter = CSVLongAdapter()
    
    headers = adapter.get_headers(responses)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Survey"
    
    # Write headers with bold font
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
    
    # Write data
    current_row = 2
    for response in responses:
        rows = adapter.export(response)
        for row_data in rows:
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=current_row, column=col_idx, value=row_data.get(header, ""))
            current_row += 1
    
    wb.save(destination)
    return destination

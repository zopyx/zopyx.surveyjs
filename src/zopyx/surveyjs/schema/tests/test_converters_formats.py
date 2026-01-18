import csv
import json
import tempfile
import unittest
import xml.etree.ElementTree as ET
from pathlib import Path

from docx import Document
from openpyxl import load_workbook

from zopyx.surveyjs.converters import (
    Attachment,
    Item,
    build_markdown,
    build_table_rows,
    build_xml,
    inline_html_images,
    render_markdown_table,
    render_text_table,
    wrap_html_output,
    wrap_pdf_html,
    write_csv,
    write_docx,
    write_html,
    write_json,
    write_pdf,
    write_text,
    write_xlsx,
    write_xml,
)
from zopyx.surveyjs.converters.html import build_html


def sample_items() -> list[Item]:
    image = Attachment("photo.png", b"\x89PNG", "image/png", field_label="Upload")
    binary = Attachment(
        "notes.bin", b"\x00\x01", "application/octet-stream", field_label="Upload"
    )
    return [
        Item(key="name", label="Name", values=["Alice"], attachments=[]),
        Item(
            key="upload",
            label="Upload",
            values=["photo.png"],
            attachments=[image, binary],
        ),
        Item(
            key="table",
            label="Table",
            values=[],
            attachments=[],
            table=[["H1", "H2"], ["A", "B"]],
            table_columns=[("c1", "Col 1"), ("c2", "Col 2")],
        ),
        Item(
            key="matrix",
            label="Matrix",
            values=["fallback"],
            attachments=[],
            field_type="matrixdynamic",
            raw_value=[{"row": 1}],
        ),
    ]


class ConverterHelpersTests(unittest.TestCase):
    def test_render_text_table(self) -> None:
        table = [["A", "B"], ["1", "2"]]
        lines = render_text_table(table)
        self.assertEqual(lines[0], "A | B")
        self.assertIn("-+-", lines[1])

    def test_render_markdown_table(self) -> None:
        table = [["A", "B"], ["1", "2"]]
        lines = render_markdown_table(table)
        self.assertEqual(lines[0], "| A | B |")
        self.assertEqual(lines[1], "| --- | --- |")

    def test_build_table_rows(self) -> None:
        rows = build_table_rows(sample_items())
        self.assertEqual(rows[0][0], "name")
        self.assertIn("photo.png", rows[1][3])

    def test_inline_html_images(self) -> None:
        image = Attachment("photo.png", b"\x89PNG", "image/png")
        html = '<img src="photo.png">'
        updated = inline_html_images(html, [image])
        self.assertIn("data:image/png;base64", updated)

    def test_wrap_html_output(self) -> None:
        wrapped = wrap_html_output("<p>Hello</p>")
        self.assertIn("<html>", wrapped)
        self.assertIn("<p>Hello</p>", wrapped)

    def test_wrap_pdf_html_metadata(self) -> None:
        wrapped = wrap_pdf_html(
            "<p>Hello</p>", creator="Ada", created="2024-01-01T10:00:00Z"
        )
        self.assertIn("Created by:", wrapped)
        self.assertIn("Created on:", wrapped)


class ConverterWritersTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmpdir = tempfile.TemporaryDirectory()
        self.base = Path(self.tmpdir.name)

    def tearDown(self) -> None:
        self.tmpdir.cleanup()

    def test_text_converter(self) -> None:
        dest = self.base / "out.txt"
        write_text(sample_items(), dest)
        content = dest.read_text(encoding="utf-8")
        self.assertIn("Survey response", content)
        self.assertIn("Attachment: photo.png", content)

    def test_markdown_and_html_converter(self) -> None:
        poll_id = "poll-1"
        md = build_markdown(sample_items(), poll_id)
        self.assertIn("# Survey response (poll-1)", md)
        self.assertIn("![Upload - photo.png]", md)
        html_body = build_html(md, sample_items()[1].attachments)
        self.assertIn("<h1>", html_body)

        dest = self.base / "out.html"
        write_html(md, sample_items()[1].attachments, dest)
        wrapped = dest.read_text(encoding="utf-8")
        self.assertIn("<html>", wrapped)
        self.assertIn("data:image/png;base64", wrapped)

    def test_pdf_converter(self) -> None:
        dest = self.base / "out.pdf"
        write_pdf(
            "<h1>Survey</h1>", dest, creator="Ada", created="2024-01-01T10:00:00Z"
        )
        data = dest.read_bytes()
        self.assertTrue(data.startswith(b"%PDF"))

    def test_csv_converter(self) -> None:
        rows = build_table_rows(sample_items())
        dest = self.base / "out.csv"
        write_csv(rows, dest)
        with dest.open("r", encoding="utf-8") as fh:
            reader = csv.reader(fh)
            header = next(reader)
            self.assertEqual(header, ["Key", "Field", "Value", "Attachments"])
            first = next(reader)
            self.assertEqual(first[0], "name")

    def test_xlsx_converter(self) -> None:
        rows = build_table_rows(sample_items())
        dest = self.base / "out.xlsx"
        write_xlsx(rows, dest)
        wb = load_workbook(dest)
        ws = wb.active
        self.assertEqual(ws.title, "Survey")
        self.assertEqual(ws["A1"].value, "Key")
        self.assertEqual(ws["A2"].value, "name")

    def test_xml_converter(self) -> None:
        dest = self.base / "out.xml"
        xml_payload = build_xml(sample_items(), "poll-1")
        self.assertIn('poll_id="poll-1"', xml_payload)
        write_xml(sample_items(), "poll-1", dest)

        root = ET.fromstring(dest.read_text(encoding="utf-8").split("\n", 1)[1])
        self.assertEqual(root.tag, "survey_response")
        fields = root.findall("field")
        self.assertEqual(len(fields), 4)
        table = fields[2].find("table")
        self.assertIsNotNone(table)
        cell = table.find("row/cell")
        self.assertEqual(cell.get("label"), "Col 1")

    def test_docx_converter(self) -> None:
        dest = self.base / "out.docx"
        write_docx(
            sample_items(),
            dest,
            poll_id="poll-1",
            creator="Ada",
            created="2024-01-01T10:00:00Z",
        )
        doc = Document(dest)
        text = "\n".join(p.text for p in doc.paragraphs)
        self.assertIn("Survey Response: poll-1", text)
        self.assertIn("Attachment: photo.png", text)

    def test_json_converter(self) -> None:
        dest = self.base / "out.json"
        write_json(
            sample_items(),
            "poll-1",
            dest,
            creator="Ada",
            created="2024-01-01T10:00:00Z",
        )
        payload = json.loads(dest.read_text(encoding="utf-8"))
        self.assertEqual(payload["poll_id"], "poll-1")
        self.assertEqual(payload["creator"], "Ada")
        self.assertEqual(payload["fields"][0]["key"], "name")
        matrix_values = payload["fields"][3]["values"]
        self.assertEqual(matrix_values, [{"row": 1}])


if __name__ == "__main__":
    unittest.main()

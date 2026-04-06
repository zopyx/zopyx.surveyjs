"""Lossless tabular export helpers for SurveyJS responses.

The exporter keeps a canonical typed representation of a response and produces
an Excel/CSV-friendly bundle with these related tables:

- responses_wide
- answers_long
- attachments
- schema
"""

from __future__ import annotations

import ast
import csv
import json
import mimetypes
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

METADATA_COLUMNS = [
    "response_id",
    "survey_id",
    "submitted_at",
    "submitted_by",
]
LONG_COLUMNS = [
    "response_id",
    "survey_id",
    "submitted_at",
    "submitted_by",
    "question_key",
    "question_title",
    "question_type",
    "path",
    "repeat_index",
    "item_index",
    "row_key",
    "row_label",
    "column_key",
    "column_label",
    "value_type",
    "value_json",
    "display_value",
]
ATTACHMENT_COLUMNS = [
    "response_id",
    "survey_id",
    "submitted_at",
    "submitted_by",
    "question_key",
    "question_title",
    "question_type",
    "item_index",
    "asset_id",
    "filename",
    "mime_type",
    "size_bytes",
    "storage_path",
    "sha256",
    "kind",
]
SCHEMA_COLUMNS = [
    "question_key",
    "question_title",
    "question_type",
    "column_name",
    "path_pattern",
    "choice_value",
    "choice_label",
    "row_key",
    "row_label",
    "column_key",
    "column_label",
]
DISPLAY_ONLY_TYPES = {"html"}
MULTI_VALUE_TYPES = {"checkbox", "tagbox", "ranking"}
ATTACHMENT_TYPES = {"file", "signaturepad"}
DYNAMIC_TYPES = {"paneldynamic", "matrixdynamic"}
RESERVED_TOP_LEVEL_KEYS = {
    "answers",
    "attachments",
    "fields",
    "result",
    "response_id",
    "survey_id",
    "poll_id",
    "id",
    "creator",
    "user",
    "submitted_by",
    "created",
    "submitted_at",
    "seq_no",
}


@dataclass
class ColumnSpec:
    """Nested column metadata for composite SurveyJS fields."""

    key: str
    label: str
    cell_type: str | None = None
    choices: dict[str, str] = field(default_factory=dict)

    @property
    def label_lookup(self) -> dict[str, str]:
        return {label: key for key, label in self.choices.items()}


@dataclass
class QuestionSpec:
    """Normalized metadata for one SurveyJS question."""

    key: str
    title: str
    question_type: str
    input_type: str | None = None
    choices: dict[str, str] = field(default_factory=dict)
    rows: dict[str, str] = field(default_factory=dict)
    columns: dict[str, ColumnSpec] = field(default_factory=dict)
    items: dict[str, str] = field(default_factory=dict)
    template_fields: dict[str, str] = field(default_factory=dict)

    @property
    def choice_lookup(self) -> dict[str, str]:
        return {label: key for key, label in self.choices.items()}

    @property
    def row_lookup(self) -> dict[str, str]:
        return {label: key for key, label in self.rows.items()}

    def choice_label(self, value: Any) -> str:
        token = "" if value is None else str(value)
        return self.choices.get(token, token)

    def resolve_choice_value(self, value: Any) -> Any:
        token = "" if value is None else str(value)
        if token in self.choices:
            return token
        return self.choice_lookup.get(token, value)

    def row_label(self, key: Any) -> str:
        token = "" if key is None else str(key)
        return self.rows.get(token, token)

    def resolve_row_key(self, value: Any) -> Any:
        token = "" if value is None else str(value)
        if token in self.rows:
            return token
        return self.row_lookup.get(token, value)


@dataclass
class AttachmentRecord:
    """Reference metadata for file-like values."""

    question_key: str
    question_title: str
    question_type: str
    item_index: int
    asset_id: str
    filename: str
    mime_type: str | None
    kind: str
    size_bytes: int | None = None
    storage_path: str | None = None
    sha256: str | None = None

    def to_row(self, response: "CanonicalResponse") -> dict[str, Any]:
        return {
            "response_id": response.response_id,
            "survey_id": response.survey_id,
            "submitted_at": response.submitted_at,
            "submitted_by": response.submitted_by,
            "question_key": self.question_key,
            "question_title": self.question_title,
            "question_type": self.question_type,
            "item_index": self.item_index,
            "asset_id": self.asset_id,
            "filename": self.filename,
            "mime_type": self.mime_type,
            "size_bytes": self.size_bytes,
            "storage_path": self.storage_path,
            "sha256": self.sha256,
            "kind": self.kind,
        }

    def to_reference(self) -> dict[str, Any]:
        return {
            "asset_id": self.asset_id,
            "filename": self.filename,
            "mime_type": self.mime_type,
            "kind": self.kind,
        }


@dataclass
class CanonicalResponse:
    """Typed, lossless response representation."""

    response_id: str
    survey_id: str
    submitted_at: str | None
    submitted_by: str | None
    answers: dict[str, Any]
    attachments: dict[str, list[AttachmentRecord]]

    def to_dict(self) -> dict[str, Any]:
        return {
            "response_id": self.response_id,
            "survey_id": self.survey_id,
            "submitted_at": self.submitted_at,
            "submitted_by": self.submitted_by,
            "answers": self.answers,
            "attachments": {
                key: [attachment.to_reference() for attachment in values]
                for key, values in self.attachments.items()
            },
        }


@dataclass
class ExportSheet:
    """A named tabular export sheet."""

    name: str
    columns: list[str]
    rows: list[dict[str, Any]]


@dataclass
class TabularExportBundle:
    """Complete export payload for CSV/XLSX output."""

    canonical_response: CanonicalResponse
    sheets: list[ExportSheet]

    def sheet(self, name: str) -> ExportSheet:
        for sheet in self.sheets:
            if sheet.name == name:
                return sheet
        raise KeyError(name)


class SurveySchemaIndex:
    """Index SurveyJS schema elements and preserve field order."""

    def __init__(self, form_payload: dict[str, Any]) -> None:
        self.form_payload = form_payload
        self.specs: dict[str, QuestionSpec] = {}
        self.order: list[str] = []
        for page in form_payload.get("pages", []):
            self._walk_elements(page.get("elements", []))

    def _walk_elements(self, elements: Iterable[dict[str, Any]]) -> None:
        for element in elements:
            if not isinstance(element, dict):
                continue
            element_type = element.get("type")
            if element_type == "panel":
                self._walk_elements(element.get("elements", []))
                continue
            name = element.get("name")
            if not name:
                continue
            self.specs[name] = QuestionSpec(
                key=name,
                title=str(element.get("title") or name),
                question_type=str(element_type or "unknown"),
                input_type=element.get("inputType"),
                choices=_choice_map(element.get("choices", [])),
                rows=_choice_map(element.get("rows", [])),
                columns=_column_map(element.get("columns", [])),
                items=_item_map(element.get("items", [])),
                template_fields=_item_map(element.get("templateElements", [])),
            )
            self.order.append(name)

    def get(self, key: str) -> QuestionSpec | None:
        return self.specs.get(key)


def _choice_map(items: Iterable[Any]) -> dict[str, str]:
    mapped: dict[str, str] = {}
    for item in items or []:
        if isinstance(item, dict):
            value = item.get("value", item.get("name"))
            label = item.get("text", item.get("title", value))
        else:
            value = item
            label = item
        if value is None:
            continue
        mapped[str(value)] = str(label if label is not None else value)
    return mapped


def _column_map(columns: Iterable[dict[str, Any]]) -> dict[str, ColumnSpec]:
    mapped: dict[str, ColumnSpec] = {}
    for column in columns or []:
        if not isinstance(column, dict):
            continue
        key = column.get("name", column.get("value"))
        if key is None:
            continue
        key_str = str(key)
        mapped[key_str] = ColumnSpec(
            key=key_str,
            label=str(column.get("title") or key_str),
            cell_type=column.get("cellType"),
            choices=_choice_map(column.get("choices", [])),
        )
    return mapped


def _item_map(items: Iterable[dict[str, Any]]) -> dict[str, str]:
    mapped: dict[str, str] = {}
    for item in items or []:
        if not isinstance(item, dict):
            continue
        key = item.get("name")
        if key is None:
            continue
        mapped[str(key)] = str(item.get("title") or key)
    return mapped


def load_json_document(path: str | Path) -> dict[str, Any]:
    """Load a JSON object from disk."""
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise TypeError(f"Expected JSON object in {path!s}")
    return payload


def build_tabular_export(
    form_payload: dict[str, Any],
    result_payload: dict[str, Any],
    survey_id: str | None = None,
) -> TabularExportBundle:
    """Build a canonical response and related tabular sheets."""
    schema = SurveySchemaIndex(form_payload)
    canonical = build_canonical_response(
        form_payload,
        result_payload,
        survey_id=survey_id,
    )
    return TabularExportBundle(
        canonical_response=canonical,
        sheets=[
            ExportSheet(
                name="responses_wide",
                columns=_wide_columns(schema, canonical),
                rows=[_wide_row(schema, canonical)],
            ),
            ExportSheet(
                name="answers_long",
                columns=LONG_COLUMNS,
                rows=_long_rows(schema, canonical),
            ),
            ExportSheet(
                name="attachments",
                columns=ATTACHMENT_COLUMNS,
                rows=_attachment_rows(canonical),
            ),
            ExportSheet(
                name="schema",
                columns=SCHEMA_COLUMNS,
                rows=_schema_rows(schema),
            ),
        ],
    )


def build_canonical_response(
    form_payload: dict[str, Any],
    result_payload: dict[str, Any],
    survey_id: str | None = None,
) -> CanonicalResponse:
    """Normalize raw or legacy result payloads into the canonical format."""
    schema = SurveySchemaIndex(form_payload)
    meta = _extract_metadata(result_payload, survey_id=survey_id)

    if isinstance(result_payload.get("answers"), dict):
        answers = _normalize_raw_answers(schema, result_payload["answers"])
        attachments = _normalize_attachment_map(
            schema,
            result_payload.get("attachments") or {},
        )
    elif isinstance(result_payload.get("fields"), list):
        answers, attachments = _normalize_legacy_fields(
            schema,
            result_payload.get("fields") or [],
        )
    else:
        raw_entry = result_payload.get("result")
        if isinstance(raw_entry, dict):
            raw_answers = raw_entry
        else:
            raw_answers = {
                key: value
                for key, value in result_payload.items()
                if key not in RESERVED_TOP_LEVEL_KEYS
            }
        answers, attachments = _normalize_raw_entry(schema, raw_answers)

    return CanonicalResponse(
        response_id=meta["response_id"],
        survey_id=meta["survey_id"],
        submitted_at=meta["submitted_at"],
        submitted_by=meta["submitted_by"],
        answers=answers,
        attachments=attachments,
    )


def write_canonical_json(bundle: TabularExportBundle, destination: str | Path) -> Path:
    """Persist the canonical typed response as JSON."""
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(
        json.dumps(bundle.canonical_response.to_dict(), ensure_ascii=False, indent=2)
        + "\n",
        encoding="utf-8",
    )
    return destination


def write_csv_bundle(bundle: TabularExportBundle, destination_dir: str | Path) -> Path:
    """Write each export sheet to a standalone UTF-8 CSV file."""
    destination_dir = Path(destination_dir)
    destination_dir.mkdir(parents=True, exist_ok=True)
    for sheet in bundle.sheets:
        path = destination_dir / f"{sheet.name}.csv"
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=sheet.columns)
            writer.writeheader()
            for row in sheet.rows:
                writer.writerow(
                    {
                        column: _csv_value(row.get(column))
                        for column in sheet.columns
                    }
                )
    return destination_dir


def write_excel_bundle(bundle: TabularExportBundle, destination: str | Path) -> Path:
    """Write the export bundle as a multi-sheet Excel workbook."""
    destination = Path(destination)
    workbook = Workbook()
    workbook.remove(workbook.active)

    header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    header_font = Font(bold=True)

    for sheet in bundle.sheets:
        worksheet = workbook.create_sheet(title=sheet.name)
        worksheet.append(sheet.columns)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(sheet.columns))}1"
        for row in sheet.rows:
            worksheet.append(
                [_excel_value(row.get(column)) for column in sheet.columns]
            )
        _autosize_columns(worksheet)

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def _extract_metadata(
    result_payload: dict[str, Any],
    survey_id: str | None = None,
) -> dict[str, str | None]:
    result_entry = result_payload.get("result")
    response_id = (
        result_payload.get("response_id")
        or result_payload.get("poll_id")
        or result_payload.get("id")
        or (result_entry or {}).get("uuid")
        or "response"
    )
    resolved_survey_id = survey_id or result_payload.get("survey_id") or "survey"
    return {
        "response_id": str(response_id),
        "survey_id": str(resolved_survey_id),
        "submitted_at": _string_or_none(
            result_payload.get("submitted_at") or result_payload.get("created")
        ),
        "submitted_by": _string_or_none(
            result_payload.get("submitted_by")
            or result_payload.get("creator")
            or result_payload.get("user")
        ),
    }


def _string_or_none(value: Any) -> str | None:
    if value is None:
        return None
    return str(value)


def _normalize_attachment_map(
    schema: SurveySchemaIndex,
    attachment_payload: dict[str, Any],
) -> dict[str, list[AttachmentRecord]]:
    attachments: dict[str, list[AttachmentRecord]] = {}
    for question_key, values in attachment_payload.items():
        spec = schema.get(question_key) or QuestionSpec(
            key=question_key,
            title=question_key,
            question_type="file",
        )
        iterable = values if isinstance(values, list) else [values]
        records: list[AttachmentRecord] = []
        for index, item in enumerate(iterable, start=1):
            if not isinstance(item, dict):
                continue
            filename = str(
                item.get("filename")
                or item.get("name")
                or f"{question_key}_{index}"
            )
            records.append(
                AttachmentRecord(
                    question_key=question_key,
                    question_title=spec.title,
                    question_type=spec.question_type,
                    item_index=index,
                    asset_id=str(item.get("asset_id") or f"{question_key}-{index}"),
                    filename=filename,
                    mime_type=_string_or_none(
                        item.get("mime_type") or item.get("content_type")
                    ),
                    size_bytes=_int_or_none(item.get("size_bytes")),
                    storage_path=_string_or_none(item.get("storage_path")),
                    sha256=_string_or_none(item.get("sha256")),
                    kind=str(item.get("kind") or _attachment_kind(spec.question_type)),
                )
            )
        if records:
            attachments[question_key] = records
    return attachments


def _normalize_legacy_fields(
    schema: SurveySchemaIndex,
    fields: list[dict[str, Any]],
) -> tuple[dict[str, Any], dict[str, list[AttachmentRecord]]]:
    answers: dict[str, Any] = {}
    attachments: dict[str, list[AttachmentRecord]] = {}
    for field_payload in fields:
        if not isinstance(field_payload, dict):
            continue
        key = str(field_payload.get("key") or field_payload.get("name") or "")
        if not key:
            continue
        spec = schema.get(key) or QuestionSpec(
            key=key,
            title=str(field_payload.get("label") or key),
            question_type="unknown",
        )
        value, field_attachments = _normalize_legacy_field(spec, field_payload)
        if value is not _MISSING and spec.question_type not in DISPLAY_ONLY_TYPES:
            answers[key] = value
        if field_attachments:
            attachments[key] = field_attachments
    return answers, attachments


def _normalize_legacy_field(
    spec: QuestionSpec,
    field_payload: dict[str, Any],
) -> tuple[Any, list[AttachmentRecord]]:
    raw_values = field_payload.get("values") or []
    attachments_meta = field_payload.get("attachments") or []

    if spec.question_type in DISPLAY_ONLY_TYPES:
        return _MISSING, []
    if spec.question_type in ATTACHMENT_TYPES:
        return _MISSING, _legacy_attachment_records(spec, raw_values, attachments_meta)
    if spec.question_type == "matrix":
        return _parse_legacy_matrix(spec, raw_values), []
    if spec.question_type == "matrixdropdown":
        value = _parse_jsonish(_first_scalar(raw_values))
        value = value if isinstance(value, dict) else {}
        return _normalize_matrixdropdown_value(spec, value), []
    if spec.question_type == "multipletext":
        value = _parse_jsonish(_first_scalar(raw_values))
        return (value if isinstance(value, dict) else {}), []
    if spec.question_type == "paneldynamic":
        value = _parse_jsonish(_first_scalar(raw_values), wrap_sequence=True)
        if isinstance(value, dict):
            value = [value]
        return (value if isinstance(value, list) else []), []
    if spec.question_type == "matrixdynamic":
        value = _parse_jsonish(_first_scalar(raw_values))
        if isinstance(value, dict):
            value = list(value.values())
        return (value if isinstance(value, list) else []), []
    if spec.question_type in MULTI_VALUE_TYPES:
        return _split_multi_values(_first_scalar(raw_values)), []
    if spec.question_type == "boolean":
        return _parse_bool(_first_scalar(raw_values)), []
    return _coerce_scalar_value(spec, _first_scalar(raw_values)), []


def _normalize_raw_entry(
    schema: SurveySchemaIndex,
    raw_answers: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, list[AttachmentRecord]]]:
    answers = _normalize_raw_answers(schema, raw_answers)
    attachments: dict[str, list[AttachmentRecord]] = {}
    for key, value in raw_answers.items():
        spec = schema.get(key)
        if spec and spec.question_type in ATTACHMENT_TYPES:
            records = _raw_attachment_records(spec, value)
            if records:
                attachments[key] = records
    return answers, attachments


def _normalize_raw_answers(
    schema: SurveySchemaIndex,
    raw_answers: dict[str, Any],
) -> dict[str, Any]:
    answers: dict[str, Any] = {}
    for key, value in raw_answers.items():
        spec = schema.get(key)
        if spec and spec.question_type in DISPLAY_ONLY_TYPES | ATTACHMENT_TYPES:
            continue
        if spec is None:
            if key == "uuid":
                continue
            answers[key] = value
            continue
        normalized = _normalize_raw_value(spec, value)
        if normalized is not _MISSING:
            answers[key] = normalized
    return answers


def _normalize_raw_value(spec: QuestionSpec, value: Any) -> Any:
    if spec.question_type in DISPLAY_ONLY_TYPES | ATTACHMENT_TYPES:
        return _MISSING
    if spec.question_type == "matrix":
        if not isinstance(value, dict):
            return {}
        return {
            str(row_key): spec.resolve_choice_value(cell_value)
            for row_key, cell_value in value.items()
        }
    if spec.question_type == "matrixdropdown":
        if not isinstance(value, dict):
            return {}
        return _normalize_matrixdropdown_value(spec, value)
    if spec.question_type == "multipletext":
        return value if isinstance(value, dict) else {}
    if spec.question_type == "paneldynamic":
        return value if isinstance(value, list) else []
    if spec.question_type == "matrixdynamic":
        if isinstance(value, list):
            return value
        if isinstance(value, dict):
            return list(value.values())
        return []
    if spec.question_type in MULTI_VALUE_TYPES:
        return value if isinstance(value, list) else _split_multi_values(value)
    if spec.question_type == "boolean":
        return value if isinstance(value, bool) else _parse_bool(value)
    return _coerce_scalar_value(spec, value)


def _parse_legacy_matrix(spec: QuestionSpec, raw_values: list[Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for raw in raw_values:
        if not isinstance(raw, str) or ":" not in raw:
            continue
        row_token, choice_token = raw.split(":", 1)
        row_key = str(spec.resolve_row_key(row_token.strip()))
        result[row_key] = spec.resolve_choice_value(choice_token.strip())
    return result


def _normalize_matrixdropdown_value(
    spec: QuestionSpec,
    value: dict[str, Any],
) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for row_key, row_value in value.items():
        resolved_row_key = str(spec.resolve_row_key(row_key))
        if not isinstance(row_value, dict):
            normalized[resolved_row_key] = row_value
            continue
        normalized_row: dict[str, Any] = {}
        for column_key, cell_value in row_value.items():
            column_token = str(column_key)
            column_spec = spec.columns.get(column_token)
            if column_spec and column_spec.choices:
                token = str(cell_value)
                if token in column_spec.choices:
                    normalized_row[column_token] = token
                else:
                    normalized_row[column_token] = column_spec.label_lookup.get(
                        token,
                        cell_value,
                    )
            else:
                normalized_row[column_token] = cell_value
        normalized[resolved_row_key] = normalized_row
    return normalized


def _legacy_attachment_records(
    spec: QuestionSpec,
    raw_values: list[Any],
    attachments_meta: list[dict[str, Any]],
) -> list[AttachmentRecord]:
    records: list[AttachmentRecord] = []
    for index, item in enumerate(attachments_meta or [], start=1):
        if not isinstance(item, dict):
            continue
        filename = str(
            item.get("filename")
            or item.get("name")
            or f"{spec.key}_{index}"
        )
        records.append(
            AttachmentRecord(
                question_key=spec.key,
                question_title=spec.title,
                question_type=spec.question_type,
                item_index=index,
                asset_id=f"{spec.key}-{index}",
                filename=filename,
                mime_type=_string_or_none(
                    item.get("mime_type") or item.get("content_type")
                ),
                kind=_attachment_kind(spec.question_type),
            )
        )
    if records:
        return records

    if spec.question_type == "signaturepad":
        signature_value = _first_scalar(raw_values)
        if isinstance(signature_value, str) and signature_value.startswith("data:"):
            mime_type, _ = _parse_data_url(signature_value)
            ext = mimetypes.guess_extension(mime_type or "image/png") or ".png"
            return [
                AttachmentRecord(
                    question_key=spec.key,
                    question_title=spec.title,
                    question_type=spec.question_type,
                    item_index=1,
                    asset_id=f"{spec.key}-1",
                    filename=f"{spec.key}{ext}",
                    mime_type=mime_type or "image/png",
                    kind="signature",
                )
            ]

    if spec.question_type == "file":
        value = _first_scalar(raw_values)
        if isinstance(value, str) and value.startswith("stored attachment:"):
            filename = value.split(":", 1)[1].strip() or f"{spec.key}-1"
            return [
                AttachmentRecord(
                    question_key=spec.key,
                    question_title=spec.title,
                    question_type=spec.question_type,
                    item_index=1,
                    asset_id=f"{spec.key}-1",
                    filename=filename,
                    mime_type=mimetypes.guess_type(filename)[0],
                    kind="file",
                )
            ]
    return []


def _raw_attachment_records(spec: QuestionSpec, value: Any) -> list[AttachmentRecord]:
    items = value if isinstance(value, list) else [value]
    records: list[AttachmentRecord] = []
    for index, item in enumerate(items, start=1):
        if isinstance(item, dict):
            content = item.get("content") or item.get("base64")
            mime_type = _string_or_none(item.get("type"))
            filename = str(
                item.get("name")
                or item.get("filename")
                or f"{spec.key}_{index}"
            )
            size_bytes = _payload_size_bytes(content)
        elif isinstance(item, str):
            content = item
            mime_type, _ = _parse_data_url(item)
            ext = mimetypes.guess_extension(
                mime_type or "application/octet-stream"
            ) or ".bin"
            filename = f"{spec.key}_{index}{ext}"
            size_bytes = _payload_size_bytes(content)
        else:
            continue
        records.append(
            AttachmentRecord(
                question_key=spec.key,
                question_title=spec.title,
                question_type=spec.question_type,
                item_index=index,
                asset_id=f"{spec.key}-{index}",
                filename=filename,
                mime_type=mime_type,
                size_bytes=size_bytes,
                kind=_attachment_kind(spec.question_type),
            )
        )
    return records


def _attachment_kind(question_type: str) -> str:
    return "signature" if question_type == "signaturepad" else "file"


def _parse_data_url(value: str) -> tuple[str | None, str]:
    if not value.startswith("data:") or "," not in value:
        return None, value
    header, payload = value.split(",", 1)
    meta = header[5:]
    mime_type = meta.split(";", 1)[0] if ";" in meta else meta
    return mime_type or None, payload


def _payload_size_bytes(value: Any) -> int | None:
    if not isinstance(value, str):
        return None
    _, payload = _parse_data_url(value)
    payload = payload.strip()
    if not payload:
        return None
    return (len(payload) * 3) // 4


def _parse_jsonish(value: Any, wrap_sequence: bool = False) -> Any:
    if not isinstance(value, str):
        return value
    stripped = value.strip()
    if not stripped:
        return None
    candidates = [stripped]
    if wrap_sequence and not stripped.startswith("["):
        candidates.append(f"[{stripped}]")
    for candidate in candidates:
        try:
            return json.loads(candidate)
        except Exception:
            pass
        try:
            return ast.literal_eval(candidate)
        except Exception:
            pass
    return value


def _parse_bool(value: Any) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    token = str(value).strip().lower()
    if token in {"yes", "true", "1", "on"}:
        return True
    if token in {"no", "false", "0", "off"}:
        return False
    return None


def _coerce_scalar_value(spec: QuestionSpec, value: Any) -> Any:
    if value is None:
        return None
    if spec.choices and isinstance(value, str):
        return spec.resolve_choice_value(value)
    if isinstance(value, (bool, int, float)):
        return value
    if isinstance(value, str) and (spec.question_type == "rating" or spec.input_type == "number"):
        number = _parse_number(value)
        return number if number is not None else value
    return value


def _parse_number(value: str) -> int | float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number.is_integer():
        return int(number)
    return number


def _split_multi_values(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if not isinstance(value, str):
        return [value]
    return [part.strip() for part in value.split(",") if part.strip()]


def _first_scalar(values: list[Any]) -> Any:
    if not values:
        return None
    return values[0]


def _int_or_none(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _wide_columns(schema: SurveySchemaIndex, response: CanonicalResponse) -> list[str]:
    columns = METADATA_COLUMNS[:]
    seen = set(columns)
    ordered_keys = list(schema.order)
    for key in response.answers:
        if key not in schema.specs:
            ordered_keys.append(key)

    for key in ordered_keys:
        value = response.answers.get(key, _MISSING)
        if value is _MISSING:
            continue
        for column_name in _question_wide_columns(schema.get(key), key, value):
            if column_name not in seen:
                columns.append(column_name)
                seen.add(column_name)
    return columns


def _question_wide_columns(
    spec: QuestionSpec | None,
    key: str,
    value: Any,
) -> list[str]:
    if spec is None:
        if isinstance(value, dict) and all(not isinstance(v, (dict, list)) for v in value.values()):
            return [f"{key}__{subkey}" for subkey in value.keys()]
        if isinstance(value, list):
            return []
        return [key]

    if spec.question_type in DISPLAY_ONLY_TYPES | MULTI_VALUE_TYPES | ATTACHMENT_TYPES | DYNAMIC_TYPES:
        return []
    if spec.question_type == "matrix" and isinstance(value, dict):
        row_keys = list(spec.rows.keys())
        extras = [row_key for row_key in value.keys() if row_key not in row_keys]
        return [f"{key}__{row_key}" for row_key in [*row_keys, *extras]]
    if spec.question_type == "multipletext" and isinstance(value, dict):
        item_keys = list(spec.items.keys())
        extras = [item_key for item_key in value.keys() if item_key not in item_keys]
        return [f"{key}__{item_key}" for item_key in [*item_keys, *extras]]
    if spec.question_type == "matrixdropdown" and isinstance(value, dict):
        column_names: list[str] = []
        row_keys = list(spec.rows.keys())
        extra_rows = [row_key for row_key in value.keys() if row_key not in row_keys]
        for row_key in [*row_keys, *extra_rows]:
            row_value = value.get(row_key, {})
            known_columns = list(spec.columns.keys())
            extra_columns = [
                column_key
                for column_key in row_value.keys()
                if column_key not in known_columns
            ] if isinstance(row_value, dict) else []
            for column_key in [*known_columns, *extra_columns]:
                column_names.append(f"{key}__{row_key}__{column_key}")
        return column_names
    if isinstance(value, dict):
        return []
    return [key]


def _wide_row(schema: SurveySchemaIndex, response: CanonicalResponse) -> dict[str, Any]:
    row: dict[str, Any] = {
        "response_id": response.response_id,
        "survey_id": response.survey_id,
        "submitted_at": response.submitted_at,
        "submitted_by": response.submitted_by,
    }
    for key, value in response.answers.items():
        _inject_wide_value(row, schema.get(key), key, value)
    return row


def _inject_wide_value(
    row: dict[str, Any],
    spec: QuestionSpec | None,
    key: str,
    value: Any,
) -> None:
    if spec is None:
        if isinstance(value, dict) and all(not isinstance(v, (dict, list)) for v in value.values()):
            for subkey, subvalue in value.items():
                row[f"{key}__{subkey}"] = subvalue
        elif not isinstance(value, list):
            row[key] = value
        return

    if spec.question_type in DISPLAY_ONLY_TYPES | MULTI_VALUE_TYPES | ATTACHMENT_TYPES | DYNAMIC_TYPES:
        return
    if spec.question_type == "matrix" and isinstance(value, dict):
        for row_key, cell_value in value.items():
            row[f"{key}__{row_key}"] = cell_value
        return
    if spec.question_type == "multipletext" and isinstance(value, dict):
        for item_key, item_value in value.items():
            row[f"{key}__{item_key}"] = item_value
        return
    if spec.question_type == "matrixdropdown" and isinstance(value, dict):
        for row_key, row_values in value.items():
            if not isinstance(row_values, dict):
                continue
            for column_key, cell_value in row_values.items():
                row[f"{key}__{row_key}__{column_key}"] = cell_value
        return
    if not isinstance(value, (dict, list)):
        row[key] = value


def _long_rows(
    schema: SurveySchemaIndex,
    response: CanonicalResponse,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    ordered_keys = list(schema.order)
    for key in response.answers:
        if key not in schema.specs:
            ordered_keys.append(key)

    for key in ordered_keys:
        if key not in response.answers:
            continue
        spec = schema.get(key) or QuestionSpec(
            key=key,
            title=key,
            question_type="unknown",
        )
        rows.extend(_question_long_rows(response, spec, response.answers[key]))

    for key, records in response.attachments.items():
        spec = schema.get(key) or QuestionSpec(
            key=key,
            title=key,
            question_type="file",
        )
        for record in records:
            rows.append(
                _base_long_row(response, spec)
                | {
                    "path": f"{key}[{record.item_index}]",
                    "item_index": record.item_index,
                    "value_type": "attachment_ref",
                    "value_json": json.dumps(
                        record.to_reference(),
                        ensure_ascii=False,
                    ),
                    "display_value": record.filename,
                }
            )
    return rows


def _question_long_rows(
    response: CanonicalResponse,
    spec: QuestionSpec,
    value: Any,
) -> list[dict[str, Any]]:
    base = _base_long_row(response, spec)
    if spec.question_type in MULTI_VALUE_TYPES and isinstance(value, list):
        rows = []
        for index, item in enumerate(value, start=1):
            rows.append(
                base
                | {
                    "path": f"{spec.key}[{index}]",
                    "item_index": index,
                    "value_type": _value_type(item),
                    "value_json": _value_json(item),
                    "display_value": spec.choice_label(item),
                }
            )
        return rows
    if spec.question_type == "matrix" and isinstance(value, dict):
        rows = []
        for row_key, cell_value in value.items():
            rows.append(
                base
                | {
                    "path": f"{spec.key}.{row_key}",
                    "row_key": row_key,
                    "row_label": spec.row_label(row_key),
                    "value_type": _value_type(cell_value),
                    "value_json": _value_json(cell_value),
                    "display_value": spec.choice_label(cell_value),
                }
            )
        return rows
    if spec.question_type == "matrixdropdown" and isinstance(value, dict):
        rows = []
        for row_key, row_values in value.items():
            if not isinstance(row_values, dict):
                continue
            for column_key, cell_value in row_values.items():
                column_spec = spec.columns.get(str(column_key))
                rows.append(
                    base
                    | {
                        "path": f"{spec.key}.{row_key}.{column_key}",
                        "row_key": row_key,
                        "row_label": spec.row_label(row_key),
                        "column_key": column_key,
                        "column_label": column_spec.label if column_spec else str(column_key),
                        "value_type": _value_type(cell_value),
                        "value_json": _value_json(cell_value),
                        "display_value": _display_nested_value(column_spec, cell_value),
                    }
                )
        return rows
    if spec.question_type == "multipletext" and isinstance(value, dict):
        rows = []
        for column_key, cell_value in value.items():
            rows.append(
                base
                | {
                    "path": f"{spec.key}.{column_key}",
                    "column_key": column_key,
                    "column_label": spec.items.get(str(column_key), str(column_key)),
                    "value_type": _value_type(cell_value),
                    "value_json": _value_json(cell_value),
                    "display_value": _display_value(cell_value),
                }
            )
        return rows
    if spec.question_type == "paneldynamic" and isinstance(value, list):
        rows = []
        for repeat_index, panel_value in enumerate(value, start=1):
            if not isinstance(panel_value, dict):
                rows.append(
                    base
                    | {
                        "path": f"{spec.key}[{repeat_index}]",
                        "repeat_index": repeat_index,
                        "value_type": _value_type(panel_value),
                        "value_json": _value_json(panel_value),
                        "display_value": _display_value(panel_value),
                    }
                )
                continue
            for column_key, cell_value in panel_value.items():
                rows.append(
                    base
                    | {
                        "path": f"{spec.key}[{repeat_index}].{column_key}",
                        "repeat_index": repeat_index,
                        "column_key": column_key,
                        "column_label": spec.template_fields.get(
                            str(column_key),
                            str(column_key),
                        ),
                        "value_type": _value_type(cell_value),
                        "value_json": _value_json(cell_value),
                        "display_value": _display_value(cell_value),
                    }
                )
        return rows
    if spec.question_type == "matrixdynamic" and isinstance(value, list):
        rows = []
        for repeat_index, row_value in enumerate(value, start=1):
            if not isinstance(row_value, dict):
                rows.append(
                    base
                    | {
                        "path": f"{spec.key}[{repeat_index}]",
                        "repeat_index": repeat_index,
                        "value_type": _value_type(row_value),
                        "value_json": _value_json(row_value),
                        "display_value": _display_value(row_value),
                    }
                )
                continue
            for column_key, cell_value in row_value.items():
                column_spec = spec.columns.get(str(column_key))
                rows.append(
                    base
                    | {
                        "path": f"{spec.key}[{repeat_index}].{column_key}",
                        "repeat_index": repeat_index,
                        "column_key": column_key,
                        "column_label": column_spec.label if column_spec else str(column_key),
                        "value_type": _value_type(cell_value),
                        "value_json": _value_json(cell_value),
                        "display_value": _display_nested_value(column_spec, cell_value),
                    }
                )
        return rows
    return [
        base
        | {
            "path": spec.key,
            "value_type": _value_type(value),
            "value_json": _value_json(value),
            "display_value": _display_scalar_for_question(spec, value),
        }
    ]


def _base_long_row(
    response: CanonicalResponse,
    spec: QuestionSpec,
) -> dict[str, Any]:
    return {
        "response_id": response.response_id,
        "survey_id": response.survey_id,
        "submitted_at": response.submitted_at,
        "submitted_by": response.submitted_by,
        "question_key": spec.key,
        "question_title": spec.title,
        "question_type": spec.question_type,
        "path": None,
        "repeat_index": None,
        "item_index": None,
        "row_key": None,
        "row_label": None,
        "column_key": None,
        "column_label": None,
        "value_type": None,
        "value_json": None,
        "display_value": None,
    }


def _display_scalar_for_question(spec: QuestionSpec, value: Any) -> str:
    if spec.question_type == "boolean":
        if value is True:
            return "Yes"
        if value is False:
            return "No"
    if spec.choices:
        return spec.choice_label(value)
    return _display_value(value)


def _display_nested_value(column_spec: ColumnSpec | None, value: Any) -> str:
    if column_spec and column_spec.choices:
        return column_spec.choices.get(str(value), str(value))
    return _display_value(value)


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _attachment_rows(response: CanonicalResponse) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for records in response.attachments.values():
        rows.extend(record.to_row(response) for record in records)
    return rows


def _schema_rows(schema: SurveySchemaIndex) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for key in schema.order:
        rows.extend(_question_schema_rows(schema.specs[key]))
    return rows


def _question_schema_rows(spec: QuestionSpec) -> list[dict[str, Any]]:
    base = {
        "question_key": spec.key,
        "question_title": spec.title,
        "question_type": spec.question_type,
        "choice_value": None,
        "choice_label": None,
        "row_key": None,
        "row_label": None,
        "column_key": None,
        "column_label": None,
    }
    rows: list[dict[str, Any]] = []

    if spec.question_type in DISPLAY_ONLY_TYPES:
        rows.append(base | {"column_name": None, "path_pattern": None})
        return rows
    if spec.question_type in ATTACHMENT_TYPES:
        rows.append(base | {"column_name": None, "path_pattern": f"{spec.key}[*]"})
        return rows
    if spec.question_type in MULTI_VALUE_TYPES:
        if spec.choices:
            for choice_value, choice_label in spec.choices.items():
                rows.append(
                    base
                    | {
                        "column_name": None,
                        "path_pattern": f"{spec.key}[*]",
                        "choice_value": choice_value,
                        "choice_label": choice_label,
                    }
                )
        else:
            rows.append(base | {"column_name": None, "path_pattern": f"{spec.key}[*]"})
        return rows
    if spec.question_type == "matrix":
        for row_key, row_label in spec.rows.items() or [(None, None)]:
            rows.append(
                base
                | {
                    "column_name": f"{spec.key}__{row_key}",
                    "path_pattern": f"{spec.key}.{row_key}",
                    "row_key": row_key,
                    "row_label": row_label,
                }
            )
        return rows
    if spec.question_type == "multipletext":
        for column_key, column_label in spec.items.items() or [(None, None)]:
            rows.append(
                base
                | {
                    "column_name": f"{spec.key}__{column_key}",
                    "path_pattern": f"{spec.key}.{column_key}",
                    "column_key": column_key,
                    "column_label": column_label,
                }
            )
        return rows
    if spec.question_type == "matrixdropdown":
        for row_key, row_label in spec.rows.items() or [(None, None)]:
            for column_key, column_spec in spec.columns.items() or [(None, ColumnSpec("value", "value"))]:
                if column_spec.choices:
                    for choice_value, choice_label in column_spec.choices.items():
                        rows.append(
                            base
                            | {
                                "column_name": f"{spec.key}__{row_key}__{column_key}",
                                "path_pattern": f"{spec.key}.{row_key}.{column_key}",
                                "choice_value": choice_value,
                                "choice_label": choice_label,
                                "row_key": row_key,
                                "row_label": row_label,
                                "column_key": column_key,
                                "column_label": column_spec.label,
                            }
                        )
                else:
                    rows.append(
                        base
                        | {
                            "column_name": f"{spec.key}__{row_key}__{column_key}",
                            "path_pattern": f"{spec.key}.{row_key}.{column_key}",
                            "row_key": row_key,
                            "row_label": row_label,
                            "column_key": column_key,
                            "column_label": column_spec.label,
                        }
                    )
        return rows
    if spec.question_type == "paneldynamic":
        for column_key, column_label in spec.template_fields.items() or [(None, None)]:
            rows.append(
                base
                | {
                    "column_name": None,
                    "path_pattern": f"{spec.key}[*].{column_key}",
                    "column_key": column_key,
                    "column_label": column_label,
                }
            )
        return rows
    if spec.question_type == "matrixdynamic":
        for column_key, column_spec in spec.columns.items() or [(None, ColumnSpec("value", "value"))]:
            rows.append(
                base
                | {
                    "column_name": None,
                    "path_pattern": f"{spec.key}[*].{column_key}",
                    "column_key": column_key,
                    "column_label": column_spec.label,
                }
            )
        return rows
    if spec.choices:
        for choice_value, choice_label in spec.choices.items():
            rows.append(
                base
                | {
                    "column_name": spec.key,
                    "path_pattern": spec.key,
                    "choice_value": choice_value,
                    "choice_label": choice_label,
                }
            )
        return rows
    rows.append(base | {"column_name": spec.key, "path_pattern": spec.key})
    return rows


def _value_type(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int) and not isinstance(value, bool):
        return "integer"
    if isinstance(value, float):
        return "number"
    if isinstance(value, list):
        return "array"
    if isinstance(value, dict):
        return "object"
    return "string"


def _value_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False)


def _csv_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _excel_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _autosize_columns(worksheet) -> None:
    for index, column in enumerate(
        worksheet.iter_cols(min_row=1, max_row=worksheet.max_row),
        start=1,
    ):
        max_length = 0
        for cell in column:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            max(max_length + 2, 10),
            60,
        )


class _MissingValue:
    pass


_MISSING = _MissingValue()

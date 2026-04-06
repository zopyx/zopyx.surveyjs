from __future__ import annotations

import csv
import importlib.util
import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
MODULE_PATH = REPO_ROOT / "src" / "zopyx" / "surveyjs" / "converters" / "tabular_export.py"
SPEC = importlib.util.spec_from_file_location("surveyjs_tabular_export_test", MODULE_PATH)
if SPEC is None or SPEC.loader is None:
    raise RuntimeError(f"Cannot load exporter module from {MODULE_PATH}")
tabular_export = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = tabular_export
SPEC.loader.exec_module(tabular_export)

build_canonical_response = tabular_export.build_canonical_response
build_tabular_export = tabular_export.build_tabular_export
write_csv_bundle = tabular_export.write_csv_bundle
write_excel_bundle = tabular_export.write_excel_bundle


def sample_form_payload() -> dict:
    return {
        "title": "All Field Types Demo",
        "pages": [
            {
                "name": "allFieldsPage",
                "title": "All Field Types",
                "elements": [
                    {"type": "text", "name": "singleLineText", "title": "Single Line Text"},
                    {"type": "text", "name": "emailField", "title": "Email", "inputType": "email"},
                    {"type": "text", "name": "numberField", "title": "Number", "inputType": "number"},
                    {"type": "text", "name": "dateField", "title": "Date", "inputType": "date"},
                    {"type": "comment", "name": "commentField", "title": "Comment / Multiline Text"},
                    {
                        "type": "radiogroup",
                        "name": "radioGroup",
                        "title": "Radio Group",
                        "choices": [
                            {"value": "option1", "text": "Option 1"},
                            {"value": "option2", "text": "Option 2"},
                            {"value": "option3", "text": "Option 3"},
                        ],
                    },
                    {
                        "type": "checkbox",
                        "name": "checkboxField",
                        "title": "Checkbox",
                        "choices": [
                            {"value": "a", "text": "Choice A"},
                            {"value": "b", "text": "Choice B"},
                            {"value": "c", "text": "Choice C"},
                        ],
                    },
                    {
                        "type": "dropdown",
                        "name": "dropdownField",
                        "title": "Dropdown",
                        "choices": [
                            {"value": "low", "text": "Low"},
                            {"value": "medium", "text": "Medium"},
                            {"value": "high", "text": "High"},
                        ],
                    },
                    {
                        "type": "tagbox",
                        "name": "tagboxField",
                        "title": "Tag Box",
                        "choices": [
                            {"value": "html", "text": "HTML"},
                            {"value": "css", "text": "CSS"},
                        ],
                    },
                    {"type": "rating", "name": "ratingField", "title": "Rating"},
                    {"type": "boolean", "name": "booleanField", "title": "Boolean"},
                    {
                        "type": "matrix",
                        "name": "matrixField",
                        "title": "Single Choice Matrix",
                        "rows": [
                            {"value": "quality", "text": "Quality"},
                            {"value": "price", "text": "Price"},
                        ],
                        "columns": [
                            {"value": "poor", "text": "Poor"},
                            {"value": "fair", "text": "Fair"},
                            {"value": "good", "text": "Good"},
                        ],
                    },
                    {
                        "type": "matrixdropdown",
                        "name": "matrixDropdownField",
                        "title": "Matrix with Dropdowns",
                        "rows": [
                            {"value": "service", "text": "Service"},
                            {"value": "product", "text": "Product"},
                        ],
                        "columns": [
                            {
                                "name": "rating",
                                "title": "Rating",
                                "cellType": "dropdown",
                                "choices": [
                                    {"value": "1", "text": "1 star"},
                                    {"value": "2", "text": "2 stars"},
                                    {"value": "3", "text": "3 stars"},
                                    {"value": "4", "text": "4 stars"},
                                    {"value": "5", "text": "5 stars"},
                                ],
                            },
                            {"name": "comment", "title": "Comment", "cellType": "text"},
                        ],
                    },
                    {
                        "type": "paneldynamic",
                        "name": "dynamicPanelField",
                        "title": "Dynamic Panel",
                        "templateElements": [
                            {"type": "text", "name": "itemName", "title": "Item Name"},
                            {"type": "text", "name": "itemValue", "title": "Value", "inputType": "number"},
                        ],
                    },
                    {"type": "file", "name": "fileUploadField", "title": "File Upload"},
                    {"type": "signaturepad", "name": "signatureField", "title": "Signature"},
                    {"type": "html", "name": "htmlContentField", "title": "HTML Content Block"},
                    {
                        "type": "imagepicker",
                        "name": "imagePickerField",
                        "title": "Image Picker",
                        "choices": [
                            {"value": "lion", "text": "Lion"},
                            {"value": "giraffe", "text": "Giraffe"},
                        ],
                    },
                    {
                        "type": "ranking",
                        "name": "rankingField",
                        "title": "Ranking",
                        "choices": [
                            {"value": "price", "text": "Price"},
                            {"value": "quality", "text": "Quality"},
                            {"value": "service", "text": "Service"},
                        ],
                    },
                ],
            }
        ],
    }


def legacy_result_payload() -> dict:
    return {
        "poll_id": "18c90d5c-2de4-11f1-8fbb-9d6e50e05304",
        "creator": "admin2",
        "created": "2026-04-01T16:01:51.681928+00:00",
        "fields": [
            {"key": "singleLineText", "label": "Single Line Text", "values": ["Sample text"], "attachments": []},
            {"key": "emailField", "label": "Email", "values": ["demo@example.com"], "attachments": []},
            {"key": "numberField", "label": "Number", "values": ["42"], "attachments": []},
            {"key": "dateField", "label": "Date", "values": ["2024-01-15"], "attachments": []},
            {"key": "commentField", "label": "Comment / Multiline Text", "values": ["This is a sample comment."], "attachments": []},
            {"key": "radioGroup", "label": "Radio Group", "values": ["option3"], "attachments": []},
            {"key": "checkboxField", "label": "Checkbox", "values": ["a, c, b"], "attachments": []},
            {"key": "dropdownField", "label": "Dropdown", "values": ["medium"], "attachments": []},
            {"key": "tagboxField", "label": "Tag Box", "values": ["html, css"], "attachments": []},
            {"key": "ratingField", "label": "Rating", "values": ["3"], "attachments": []},
            {"key": "booleanField", "label": "Boolean", "values": ["Yes"], "attachments": []},
            {"key": "matrixField", "label": "Single Choice Matrix", "values": ["Quality: Poor", "Price: Good"], "attachments": []},
            {"key": "matrixDropdownField", "label": "Matrix with Dropdowns", "values": ['{"service": {"rating": "4", "comment": "Great service"}, "product": {"rating": "5", "comment": "Excellent product"}}'], "attachments": []},
            {"key": "dynamicPanelField", "label": "Dynamic Panel", "values": ["{'itemName': 'First Item', 'itemValue': 100}, {'itemName': 'abc', 'itemValue': 5}"], "attachments": []},
            {"key": "rankingField", "label": "Ranking", "values": ["quality, service, price"], "attachments": []},
            {"key": "fileUploadField", "label": "File Upload", "values": ["stored attachment: 1004_2026-04-01_UNIV_59.pdf"], "attachments": [{"name": "1004_2026-04-01_UNIV_59.pdf", "content_type": "application/pdf", "is_image": False}]},
            {"key": "signatureField", "label": "Signature", "values": ["data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAAB"], "attachments": []},
            {"key": "imagePickerField", "label": "Image Picker", "values": ["giraffe"], "attachments": []},
        ],
    }


def test_build_canonical_response_recovers_nested_and_typed_values() -> None:
    canonical = build_canonical_response(
        sample_form_payload(),
        legacy_result_payload(),
        survey_id="survey-form-09075e2a",
    )

    assert canonical.response_id == "18c90d5c-2de4-11f1-8fbb-9d6e50e05304"
    assert canonical.survey_id == "survey-form-09075e2a"
    assert canonical.submitted_by == "admin2"
    assert canonical.answers["numberField"] == 42
    assert canonical.answers["checkboxField"] == ["a", "c", "b"]
    assert canonical.answers["matrixField"] == {"quality": "poor", "price": "good"}
    assert canonical.answers["matrixDropdownField"]["service"]["rating"] == "4"
    assert canonical.answers["dynamicPanelField"][1]["itemValue"] == 5
    assert canonical.answers["rankingField"] == ["quality", "service", "price"]
    assert "fileUploadField" not in canonical.answers
    assert canonical.attachments["fileUploadField"][0].filename == "1004_2026-04-01_UNIV_59.pdf"
    assert canonical.attachments["signatureField"][0].kind == "signature"


def test_build_tabular_export_creates_expected_wide_long_and_schema_rows() -> None:
    bundle = build_tabular_export(
        sample_form_payload(),
        legacy_result_payload(),
        survey_id="survey-form-09075e2a",
    )

    wide_sheet = bundle.sheet("responses_wide")
    wide_row = wide_sheet.rows[0]
    assert "checkboxField" not in wide_sheet.columns
    assert "matrixField__quality" in wide_sheet.columns
    assert wide_row["singleLineText"] == "Sample text"
    assert wide_row["matrixDropdownField__product__comment"] == "Excellent product"

    long_sheet = bundle.sheet("answers_long")
    long_paths = {row["path"]: row for row in long_sheet.rows}
    assert long_paths["dynamicPanelField[2].itemValue"]["value_json"] == "5"
    assert long_paths["checkboxField[2]"]["display_value"] == "Choice C"
    assert long_paths["fileUploadField[1]"]["value_type"] == "attachment_ref"

    attachment_sheet = bundle.sheet("attachments")
    assert len(attachment_sheet.rows) == 2
    assert attachment_sheet.rows[1]["kind"] == "signature"

    schema_sheet = bundle.sheet("schema")
    schema_paths = {row["path_pattern"] for row in schema_sheet.rows}
    assert "matrixDropdownField.service.rating" in schema_paths
    assert "dynamicPanelField[*].itemValue" in schema_paths


def test_write_csv_bundle_and_excel_bundle(tmp_path: Path) -> None:
    bundle = build_tabular_export(
        sample_form_payload(),
        legacy_result_payload(),
        survey_id="survey-form-09075e2a",
    )

    csv_dir = write_csv_bundle(bundle, tmp_path / "csv")
    workbook_path = write_excel_bundle(bundle, tmp_path / "survey.xlsx")

    with (csv_dir / "responses_wide.csv").open(encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    assert rows[0]["singleLineText"] == "Sample text"
    assert rows[0]["matrixField__price"] == "good"

    workbook = load_workbook(workbook_path)
    assert workbook.sheetnames == [
        "responses_wide",
        "answers_long",
        "attachments",
        "schema",
    ]
    wide_sheet = workbook["responses_wide"]
    assert wide_sheet["A1"].value == "response_id"
    headers = [cell.value for cell in wide_sheet[1]]
    product_comment_col = headers.index("matrixDropdownField__product__comment") + 1
    assert wide_sheet.cell(row=2, column=product_comment_col).value == "Excellent product"


def test_standalone_script_exports_workbook_and_optional_files(tmp_path: Path) -> None:
    form_path = tmp_path / "form.json"
    result_path = tmp_path / "result.json"
    output_path = tmp_path / "result.xlsx"
    csv_dir = tmp_path / "csv"
    canonical_path = tmp_path / "canonical.json"

    form_path.write_text(json.dumps(sample_form_payload()), encoding="utf-8")
    result_path.write_text(json.dumps(legacy_result_payload()), encoding="utf-8")

    script_path = REPO_ROOT / "scripts" / "export_survey_excel.py"
    completed = subprocess.run(
        [
            sys.executable,
            str(script_path),
            "--form",
            str(form_path),
            "--result",
            str(result_path),
            "--output",
            str(output_path),
            "--survey-id",
            "survey-form-09075e2a",
            "--csv-dir",
            str(csv_dir),
            "--canonical-json",
            str(canonical_path),
        ],
        check=True,
        capture_output=True,
        text=True,
    )

    assert "Wrote workbook" in completed.stdout
    assert output_path.exists()
    assert (csv_dir / "answers_long.csv").exists()
    canonical = json.loads(canonical_path.read_text(encoding="utf-8"))
    assert canonical["survey_id"] == "survey-form-09075e2a"
